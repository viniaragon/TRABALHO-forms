import json
import sys
from pathlib import Path

import openpyxl


def cell_to_text(value):
    if value is None:
        return None
    return str(value)


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Uso: python scripts/inspecionar-planilha-brenda.py <arquivo.xlsx>")

    workbook_path = Path(sys.argv[1])
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    output = []

    for sheet in workbook.worksheets:
        sample_rows = []
        non_empty_rows = 0
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = list(row)
            if not any(value is not None and str(value).strip() for value in values):
                continue
            non_empty_rows += 1
            if len(sample_rows) < 12:
                sample_rows.append({
                    "row": row_index,
                    "values": [cell_to_text(value) for value in values[:40]],
                })

        output.append({
            "sheet": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "non_empty_rows": non_empty_rows,
            "sample": sample_rows,
        })

    print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
